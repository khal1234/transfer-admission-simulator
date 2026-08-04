"""원본 스프레드시트에서 합격자 성적을 다시 뽑는다. 읽기 전용(파일을 쓰지 않는다).

`results/`의 JSON은 과거 Gemini CLI로 뽑은 것이고 그 파서는 리포에 남아 있지
않다. 어느 숫자가 어느 원본에서 나왔는지 되짚을 수 없어, 원본에서 독립적으로
다시 뽑는 경로를 만든다. 대조는 `tools/diff_extraction.py` 가 한다.

**원칙: 원본이 싣지 않은 값은 만들지 않는다.**
경북대-충남대는 환산점수만 공개하고 원점수를 싣지 않는다. 그런 칸의 원점수는
None 으로 둔다. 환산식을 뒤집어 채우면 그 값으로 하는 모든 검사가 순환 논증이
된다(tools/audit_circularity.py 참조).

다루는 범위는 스프레드시트 10칸이다. PDF 원본(강원-부산-인천-전남-충북,
경북 2024)은 아직 포함하지 않는다.

    python tools/extract_spreadsheets.py            # 뽑은 결과 요약
    python tools/extract_spreadsheets.py --dump     # 레코드 전문(JSON)
"""
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

# 가운뎃점 변종들. 원본과 results 가 서로 다른 문자를 쓴다
# (원본 '산림과학·조경학부' vs results '산림과학ㆍ조경학부').
MIDDLE_DOTS = "·ㆍ‧•・"


def normalize_name(value):
    """학과명 비교용 정규화. 공백-개행-가운뎃점 표기 차이를 없앤다.

    정규화 없이 비교하면 표기 차이가 전부 '누락'으로 잡힌다(실측: 가짜 누락 5건).
    """
    if value is None:
        return ""

    # 가운뎃점은 NFKC 전에 지운다. NFKC가 'ㆍ'(U+318D)를 다른 코드로 바꿔버려
    # 뒤에서 지우면 빠져나간다 — 실제로 산림과학ㆍ조경학부 3건이 그렇게
    # '원본에 없음'으로 잘못 잡혔다.
    text = str(value)
    for dot in MIDDLE_DOTS:
        text = text.replace(dot, "")

    text = unicodedata.normalize("NFKC", text)
    for dot in MIDDLE_DOTS:
        text = text.replace(dot, "")

    return re.sub(r"\s+", "", text)


def to_number(value):
    """숫자 셀만 숫자로. 빈칸-문자열-0은 값 없음으로 본다.

    0을 살려두면 '0점'으로 읽혀 평균과 비교를 오염시킨다. 원본에서 실제로
    0이 찍히는 자리는 '모집인원 없음'이나 미반영을 뜻한다(부경대 본부 소속 등).
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return None if value == 0 else float(value)

    text = str(value).strip().replace(",", "")
    if text in ("", "-", "–"):
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return None if number == 0 else number


def to_count(value):
    """인원 수. 여기서는 0이 의미 있는 값이라 살린다."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip().replace(",", "")
    if text in ("", "-", "–"):
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def read_xlsx_rows(path, sheet_name):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def read_xls_rows(path):
    import xlrd

    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    return [sheet.row_values(i) for i in range(sheet.nrows)]


def cell(row, index):
    return row[index] if index < len(row) else None


def make_record(univ, year, unit, **fields):
    record = {
        "대학명": univ,
        "연도": year,
        "학과_원본명": str(unit).strip(),
        "학과_정규화": normalize_name(unit),
        "모집인원": None,
        "지원인원": None,
        "합격인원": None,
        "최종합격_토익환산점수": None,
        "최종합격_토익원점수": None,
        "최종합격_학점환산점수": None,
        "최종합격_학점원점수_100점만점": None,
    }
    record.update(fields)
    return record


# ── 대학별 추출기 ─────────────────────────────────────────────

def extract_pukyong(year):
    """부경대: '일반편입 평균성적 현황' 시트.

    영어는 원점수(공인영어성적, TOEIC 해당자)와 환산점수
    (영어변환성적점수)를 모두 싣는다. GPA는 원 백분위가 아니라 모집요강의
    100점 배점 환산식이 적용된 '전적대학성적점수'만 싣는다.
    """
    path = DATA_DIR / f"{year[2:]}_부경대_성적.xlsx"
    rows = read_xlsx_rows(path, "일반편입 평균성적 현황")

    records = []
    for row in rows:
        unit = cell(row, 1)
        if not isinstance(unit, str) or not unit.strip():
            continue
        if unit.strip() in ("모집단위", "모집단위명"):
            continue

        records.append(make_record(
            "부경대학교", year, unit,
            모집인원=to_count(cell(row, 2)),
            지원인원=to_count(cell(row, 3)),
            합격인원=to_count(cell(row, 4)),
            최종합격_학점환산점수=to_number(cell(row, 6)),
            최종합격_토익원점수=to_number(cell(row, 8)),
            최종합격_토익환산점수=to_number(cell(row, 9)),
        ))
    return records


def find_header_columns(rows, anchor, labels, lookahead=2):
    """헤더 글자로 열 번호를 찾는다. {라벨: 열번호}.

    연도마다 열 구성이 바뀌기 때문에 번호를 박아두면 조용히 밀린다 — 실제로
    경북대 2026은 2025 대비 한 칸 밀려 172건이 전부 엉뚱한 값으로 잡혔다
    (경쟁률이 앞으로 오고 '합격자'-'최종등록자' 열이 새로 생겼다).
    """
    for i, row in enumerate(rows):
        texts = [str(c) if c is not None else "" for c in row]
        if not any(anchor in t for t in texts):
            continue

        found = {}
        for offset in range(lookahead + 1):
            if i + offset >= len(rows):
                break
            for j, c in enumerate(rows[i + offset]):
                text = re.sub(r"\s+", "", str(c)) if c is not None else ""
                for label in labels:
                    if label not in found and text.startswith(label):
                        found[label] = j
        return found

    return {}


def extract_kyungpook(year):
    """경북대: 전형구분이 병합셀이라 앞 값을 이어받아야 한다.

    이어받지 않으면 그 행이 통째로 빠진다 — 2025 생물학과가 실제로 그렇게
    results 에서 누락돼 있었다. 원본은 환산점수만 싣고 원점수는 없다.
    열 번호는 헤더 글자로 찾는다(연도마다 구성이 바뀐다).
    """
    path = DATA_DIR / f"{year[2:]}_경북대_성적.xlsx"
    rows = read_xlsx_rows(path, None)

    cols = find_header_columns(
        rows, "전형구분",
        ["모집단위", "모집인원", "지원인원", "입학인원", "최종등록자",
         "공인영어", "전적대학성적"],
    )
    unit_col = cols.get("모집단위", 2)
    # 2025는 '입학인원', 2026은 '최종등록자'로 이름이 다르다.
    enrolled_col = cols.get("최종등록자", cols.get("입학인원"))

    records = []
    admission_type = None
    for row in rows:
        # 전형구분 열에 값이 있으면 무엇이든 현재 구분으로 삼는다.
        # 아는 이름만 받으면 모르는 전형(농어촌-기회균형-특성화고교졸업자 등)이
        # 앞 구분을 그대로 물려받아 일반편입으로 섞여 들어간다.
        first = cell(row, 0)
        if isinstance(first, str) and first.strip():
            admission_type = first.strip()

        unit = cell(row, unit_col)
        if not isinstance(unit, str) or not unit.strip():
            continue
        if unit.strip() == "모집단위":
            continue
        if admission_type != "일반편입":
            continue

        records.append(make_record(
            "경북대학교", year, unit,
            모집인원=to_count(cell(row, cols["모집인원"])),
            지원인원=to_count(cell(row, cols["지원인원"])),
            합격인원=(
                to_count(cell(row, enrolled_col))
                if enrolled_col is not None else None
            ),
            최종합격_토익환산점수=to_number(cell(row, cols["공인영어"])),
            최종합격_학점환산점수=to_number(cell(row, cols["전적대학성적"])),
        ))
    return records


def extract_chungnam(year):
    """충남대: '일반편입학' 시트. 환산점수만 싣는다(원점수 열 없음).

    2024는 전적대학 성적 열이 있고, 2025부터는 그 열 자체가 사라진다
    (전형에서 빠졌다 — formulaRegistry 의 학과별 충남대 프로필과 일치).
    대학 열은 병합셀이라 이어받는다. 인원 정보는 이 시트에 없다.
    """
    path = DATA_DIR / f"{year[2:]}_충남대_성적.xlsx"
    rows = read_xlsx_rows(path, "일반편입학")

    if year == "2024":
        english_col, gpa_col = 3, 4
    else:
        english_col, gpa_col = 3, None

    records = []
    seen_header = False
    for row in rows:
        unit = cell(row, 2)
        if isinstance(unit, str) and unit.strip() == "모집단위":
            seen_header = True
            continue
        if not seen_header:
            continue
        if not isinstance(unit, str) or not unit.strip():
            continue

        records.append(make_record(
            "충남대학교", year, unit,
            최종합격_토익환산점수=to_number(cell(row, english_col)),
            최종합격_학점환산점수=(
                to_number(cell(row, gpa_col)) if gpa_col is not None else None
            ),
        ))
    return records


def extract_jeonbuk(year):
    """전북대: .xls. 원점수(전적대 백분위, 토익 평균)만 싣고 환산은 없다."""
    path = DATA_DIR / f"{year[2:]}_전북대_성적.xls"
    rows = read_xls_rows(path)

    records = []
    admission_type = None
    for row in rows:
        # 전형구분 열에 값이 있으면 무엇이든 현재 구분으로 삼는다.
        # 아는 이름만 받으면 모르는 전형(농어촌-기회균형-특성화고교졸업자 등)이
        # 앞 구분을 그대로 물려받아 일반편입으로 섞여 들어간다.
        first = cell(row, 0)
        if isinstance(first, str) and first.strip():
            admission_type = first.strip()

        unit = cell(row, 3)
        if not isinstance(unit, str) or not unit.strip():
            continue
        if unit.strip() == "모집단위":
            continue
        if admission_type != "일반편입":
            continue

        records.append(make_record(
            "전북대학교", year, unit,
            모집인원=to_count(cell(row, 4)),
            지원인원=to_count(cell(row, 5)),
            합격인원=to_count(cell(row, 6)),
            최종합격_학점원점수_100점만점=to_number(cell(row, 8)),
            최종합격_토익원점수=to_number(cell(row, 9)),
        ))
    return records


EXTRACTORS = [
    ("부경대학교", "2024", extract_pukyong),
    ("부경대학교", "2025", extract_pukyong),
    ("부경대학교", "2026", extract_pukyong),
    ("경북대학교", "2025", extract_kyungpook),
    ("경북대학교", "2026", extract_kyungpook),
    ("충남대학교", "2024", extract_chungnam),
    ("충남대학교", "2025", extract_chungnam),
    ("충남대학교", "2026", extract_chungnam),
    ("전북대학교", "2025", extract_jeonbuk),
    ("전북대학교", "2026", extract_jeonbuk),
]

# 원본이 원점수를 싣는가 — audit_circularity.py 와 같은 사실을 쓴다.
SOURCE_PUBLISHES = {
    "부경대학교": {"토익원": True, "토익환산": True, "학점원": False, "학점환산": True},
    "경북대학교": {"토익원": False, "토익환산": True, "학점원": False, "학점환산": True},
    "충남대학교": {"토익원": False, "토익환산": True, "학점원": False, "학점환산": None},
    "전북대학교": {"토익원": True, "토익환산": False, "학점원": True, "학점환산": False},
}


@dataclass(frozen=True)
class ExtractionFailure:
    university: str
    year: str
    error_type: str
    message: str


@dataclass(frozen=True)
class ExtractionResult:
    records: list[dict]
    failures: list[ExtractionFailure]
    covered: set[tuple[str, str]]


REQUIRED_RECORD_FIELDS = {
    "대학명", "연도", "학과_원본명", "학과_정규화",
    "모집인원", "지원인원", "합격인원",
    "최종합격_토익환산점수", "최종합격_토익원점수",
    "최종합격_학점환산점수", "최종합격_학점원점수_100점만점",
}


def validate_extracted_records(univ, year, records):
    if not isinstance(records, list):
        raise TypeError(f"추출기가 list 대신 {type(records).__name__} 반환")
    if not records:
        raise ValueError("원본에서 레코드를 한 행도 추출하지 못함")

    seen = set()
    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            raise TypeError(f"{index}번째 레코드가 object가 아님")
        missing = REQUIRED_RECORD_FIELDS - record.keys()
        if missing:
            raise ValueError(f"{index}번째 레코드 스키마 누락: {sorted(missing)}")
        if record["대학명"] != univ or record["연도"] != year:
            raise ValueError(f"{index}번째 레코드 대학·연도 불일치")
        normalized = record["학과_정규화"]
        if not normalized:
            raise ValueError(f"{index}번째 레코드 학과명이 비어 있음")
        if normalized in seen:
            raise ValueError(f"정규화 학과 키 중복: {record['학과_원본명']}")
        seen.add(normalized)


def extract_all(
    extractors: list[tuple[str, str, Callable[[str], list[dict]]]] | None = None,
) -> ExtractionResult:
    records = []
    failures = []
    covered = set()
    selected_extractors = EXTRACTORS if extractors is None else extractors
    for univ, year, fn in selected_extractors:
        try:
            extracted = fn(year)
            validate_extracted_records(univ, year, extracted)
        except Exception as exc:
            failures.append(ExtractionFailure(
                university=univ,
                year=year,
                error_type=exc.__class__.__name__,
                message=str(exc),
            ))
            continue
        records.extend(extracted)
        covered.add((univ, year))
    return ExtractionResult(records=records, failures=failures, covered=covered)


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    result = extract_all()
    records = result.records

    for failure in result.failures:
        print(
            f"!! {failure.year} {failure.university} 추출 실패 "
            f"[{failure.error_type}]: {failure.message}",
            file=sys.stderr,
        )

    if "--dump" in sys.argv:
        print(json.dumps(records, ensure_ascii=False, indent=1))
        return 1 if result.failures else 0

    from collections import Counter

    counts = Counter((r["대학명"], r["연도"]) for r in records)
    print("{:<10}{:<6}{:>7}{:>10}{:>10}{:>10}{:>10}".format(
        "대학", "연도", "레코드", "토익원", "토익환산", "학점원", "학점환산"))
    print("-" * 63)

    for key in sorted(counts):
        rows = [r for r in records if (r["대학명"], r["연도"]) == key]
        def filled(field):
            return sum(1 for r in rows if r[field] is not None)

        print("{:<10}{:<6}{:>7}{:>10}{:>10}{:>10}{:>10}".format(
            key[0].replace("대학교", "대"), key[1], len(rows),
            filled("최종합격_토익원점수"),
            filled("최종합격_토익환산점수"),
            filled("최종합격_학점원점수_100점만점"),
            filled("최종합격_학점환산점수"),
        ))

    print()
    print(f"합계 {len(records)}건 / {len(counts)}칸")
    if result.failures:
        print(
            f"!! {len(result.failures)}개 원본 추출 실패 — 결과를 신뢰할 수 없습니다.",
            file=sys.stderr,
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
